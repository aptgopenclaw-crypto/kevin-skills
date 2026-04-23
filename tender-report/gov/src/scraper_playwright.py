"""
政府採購網招標資料爬蟲 - Playwright 版
使用本機瀏覽器繞過 WAF + Vision AI 辨識撲克牌 CAPTCHA
"""

import asyncio
import base64
import json
import logging
import os
import re
import time
from datetime import datetime, timedelta
from urllib.parse import urlencode

import pandas as pd
from playwright.async_api import async_playwright, Page, BrowserContext

from config import KEYWORDS, HEADERS, REQUEST_TIMEOUT, OUTPUT_FILE

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Vision AI 設定 - NVIDIA API (OpenAI 相容)
VISION_API_URL = os.environ.get("VISION_API_URL", "https://integrate.api.nvidia.com/v1/chat/completions")
VISION_API_KEY = os.environ.get("VISION_API_KEY", "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")
VISION_MODEL = os.environ.get("VISION_MODEL", "meta/llama-3.2-90b-vision-instruct")

BASE_URL = "https://web.pcc.gov.tw"
SEARCH_URL = f"{BASE_URL}/prkms/tender/common/basic/readTenderBasic"


class CaptchaSolver:
    """使用 Vision AI 辨識撲克牌 CAPTCHA"""

    @staticmethod
    async def solve(screenshot_bytes: bytes) -> list[int]:
        """
        分析 CAPTCHA 截圖，回傳 B區中匹配 A區的牌的索引 (0-based)

        Returns: list of 0-based indices of matching cards in B區
        """
        import httpx

        b64_image = base64.b64encode(screenshot_bytes).decode()

        prompt = """你是一個撲克牌辨識系統。這張圖是一個 CAPTCHA 驗證碼。

圖片分為兩區：
- A區（上方）：2張目標撲克牌
- B區（下方）：6張可選撲克牌

牌面上有干擾線條和浮水印，請忽略它們，專注辨識每張牌的花色和數字。

請辨識：
1. A區的2張牌各是什麼（花色+數字）
2. B區的6張牌各是什麼（花色+數字），從左到右編號 0-5

然後找出 B區中與 A區匹配的牌（花色和數字都相同）。

請用以下 JSON 格式回答（不要包含其他文字）：
{
  "area_a": ["10♠", "A♠"],
  "area_b": ["5♥", "A♣", "8♦", "8♠", "9♦", "10♠"],
  "matches": [5]
}

matches 是 B區中匹配牌的索引（0-based）。"""

        try:
            async with httpx.AsyncClient(timeout=30) as client:
                resp = await client.post(
                    VISION_API_URL,
                    headers={
                        "Authorization": f"Bearer {VISION_API_KEY}",
                        "Content-Type": "application/json"
                    },
                    json={
                        "model": VISION_MODEL,
                        "messages": [
                            {
                                "role": "user",
                                "content": [
                                    {"type": "text", "text": prompt},
                                    {
                                        "type": "image_url",
                                        "image_url": {
                                            "url": f"data:image/png;base64,{b64_image}"
                                        }
                                    }
                                ]
                            }
                        ],
                        "max_tokens": 500,
                        "temperature": 0
                    }
                )
                resp.raise_for_status()
                result = resp.json()
                content = result["choices"][0]["message"]["content"]

                # 解析 JSON
                # 嘗試從 content 中提取 JSON
                json_match = re.search(r'\{[^}]+\}', content, re.DOTALL)
                if json_match:
                    data = json.loads(json_match.group())
                    matches = data.get("matches", [])
                    logger.info(f"CAPTCHA AI 辨識: A區={data.get('area_a')}, B區={data.get('area_b')}, 匹配={matches}")
                    return matches
                else:
                    logger.error(f"無法從 AI 回應中解析 JSON: {content}")
                    return []

        except Exception as e:
            logger.error(f"Vision AI 呼叫失敗: {e}")
            return []


class PlaywrightTenderScraper:
    def __init__(self, headless: bool = False):
        self.headless = headless
        self.data = []
        self.error_keywords = []
        self.captcha_solver = CaptchaSolver()
        self.browser = None
        self.context = None
        self.page = None

    def get_date_range(self):
        """取得當天日期範圍"""
        end_date = datetime.now()
        start_date = end_date - timedelta(days=0)
        return start_date.strftime("%Y/%m/%d"), end_date.strftime("%Y/%m/%d")

    def build_search_url(self, keyword: str) -> str:
        """建構搜尋 URL"""
        start_date, end_date = self.get_date_range()
        params = {
            'firstSearch': 'true',
            'searchType': 'basic',
            'isBinding': 'N',
            'isLogIn': 'N',
            'orgName': '',
            'orgId': '',
            'tenderName': keyword,
            'tenderId': '',
            'tenderType': 'TENDER_DECLARATION',
            'tenderWay': 'TENDER_WAY_ALL_DECLARATION',
            'dateType': 'isNow',
            'tenderStartDate': start_date,
            'tenderEndDate': end_date,
            'radProctrgCate': '',
            'policyAdvocacy': ''
        }
        return f"{SEARCH_URL}?{urlencode(params)}"

    async def start(self):
        """啟動瀏覽器"""
        self.pw = await async_playwright().start()
        self.browser = await self.pw.chromium.launch(
            headless=self.headless,
            args=['--disable-blink-features=AutomationControlled']
        )
        self.context = await self.browser.new_context(
            user_agent=HEADERS['User-Agent'],
            viewport={'width': 1280, 'height': 900},
            locale='zh-TW'
        )
        self.page = await self.context.new_page()
        logger.info(f"瀏覽器啟動完成 (headless={self.headless})")

    async def close(self):
        """關閉瀏覽器"""
        if self.browser:
            await self.browser.close()
        if self.pw:
            await self.pw.stop()

    async def solve_captcha(self) -> bool:
        """
        辨識並解決 CAPTCHA
        回傳 True 表示成功，False 表示失敗
        """
        page = self.page
        max_attempts = 3

        for attempt in range(max_attempts):
            logger.info(f"嘗試解 CAPTCHA (第 {attempt + 1}/{max_attempts} 次)")

            # 等待 CAPTCHA 頁面載入
            await page.wait_for_load_state('networkidle')
            await asyncio.sleep(1)

            # 截圖 CAPTCHA 區域
            screenshot = await page.screenshot(full_page=False)

            # 呼叫 Vision AI
            matches = await self.captcha_solver.solve(screenshot)

            if not matches:
                logger.warning("AI 未找到匹配的牌，嘗試重新整理")
                # 點重新整理按鈕
                refresh_btn = page.locator('text=重新整理').first
                if await refresh_btn.is_visible():
                    await refresh_btn.click()
                    await asyncio.sleep(2)
                continue

            # 點擊 B區匹配的牌
            # B區的牌通常是 img 或 canvas 元素，需要根據實際 DOM 定位
            # 先嘗試找 B區的所有牌
            b_cards = page.locator('#cardB img, .areaB img, [id*="cardB"] img')
            b_count = await b_cards.count()

            if b_count == 0:
                # 嘗試其他選擇器
                b_cards = page.locator('img[onclick*="selectCard"], img[onclick*="select"]')
                b_count = await b_cards.count()

            if b_count == 0:
                # 最後嘗試：找所有在 B區範圍內的可點擊圖片
                # 透過截圖座標來點擊
                logger.warning(f"無法定位 B區牌元素 (count={b_count})，嘗試座標點擊")
                # TODO: 根據截圖座標計算點擊位置
                continue

            logger.info(f"找到 B區 {b_count} 張牌")

            for idx in matches:
                if idx < b_count:
                    card = b_cards.nth(idx)
                    await card.click()
                    logger.info(f"點擊 B區第 {idx} 張牌")
                    await asyncio.sleep(0.5)

            # 點確認送出
            submit_btn = page.locator('text=確認送出').first
            if await submit_btn.is_visible():
                await submit_btn.click()
                logger.info("點擊確認送出")
                await asyncio.sleep(2)

                # 檢查是否成功（是否跳到詳細頁面）
                current_url = page.url
                if 'searchTenderDetail' in current_url:
                    logger.info("CAPTCHA 驗證成功！")
                    return True
                else:
                    # 可能驗證失敗，檢查是否還在 CAPTCHA 頁
                    page_text = await page.text_content('body')
                    if 'A區' in page_text and 'B區' in page_text:
                        logger.warning("CAPTCHA 驗證失敗，重試...")
                        continue
                    else:
                        # 可能成功了但 URL 不同
                        logger.info("CAPTCHA 可能已通過")
                        return True

        logger.error(f"CAPTCHA 解決失敗（已嘗試 {max_attempts} 次）")
        return False

    async def parse_list_page(self, keyword: str) -> list[dict]:
        """解析列表頁，回傳招標項目清單（含 detail link）"""
        page = self.page
        items = []

        await page.wait_for_load_state('networkidle')

        # 找主要資料表格
        rows = await page.query_selector_all('table.tb_01 tbody tr, table#tpam tbody tr')
        if not rows:
            rows = await page.query_selector_all('tr.tb_b2')

        if not rows:
            logger.info(f"關鍵字 '{keyword}' 未找到招標資料")
            return items

        for row in rows:
            cols = await row.query_selector_all('td')
            if len(cols) < 9:
                continue

            # 取採購性質（第5欄）
            procurement_type = await cols[5].inner_text()
            if '工程' not in procurement_type.strip():
                continue

            # 取得各欄資料
            org_name = (await cols[1].inner_text()).strip()
            col2_text = (await cols[2].inner_text()).strip()
            lines = [l.strip() for l in col2_text.split('\n') if l.strip()]
            tender_id = lines[0] if lines else ''
            tender_name = ' '.join(lines[1:]) if len(lines) > 1 else ''

            # 嘗試從 script 取標案名稱
            script_el = await cols[2].query_selector('script')
            if script_el:
                script_text = await script_el.inner_html()
                match = re.search(r'Geps3\.CNS\.pageCode2Img\("([^"]+)"\)', script_text)
                if match:
                    tender_name = match.group(1)

            transmission = (await cols[3].inner_text()).strip()
            tender_way = (await cols[4].inner_text()).strip()
            announce_date = (await cols[6].inner_text()).strip()
            deadline = (await cols[7].inner_text()).strip()
            budget = (await cols[8].inner_text()).strip()

            # 找「檢視」連結取 pkPmsMain
            view_link = await row.query_selector('a[href*="searchTenderDetail"], a[onclick*="searchTenderDetail"]')
            detail_url = None
            if view_link:
                href = await view_link.get_attribute('href')
                if href and 'pkPmsMain' in href:
                    detail_url = href if href.startswith('http') else BASE_URL + href
                else:
                    onclick = await view_link.get_attribute('onclick')
                    if onclick:
                        pk_match = re.search(r'pkPmsMain[=:]([A-Za-z0-9+/=]+)', onclick)
                        if pk_match:
                            detail_url = f"{BASE_URL}/tps/QueryTender/query/searchTenderDetail?pkPmsMain={pk_match.group(1)}"

            item = {
                '關鍵字': keyword,
                '機關名稱': org_name,
                '標案案號': tender_id,
                '標案名稱': tender_name,
                '傳輸次數': transmission,
                '招標方式': tender_way,
                '採購性質': procurement_type.strip(),
                '公告日期': announce_date,
                '截止投標日期': deadline,
                '預算金額': budget,
                '_detail_url': detail_url,
            }
            items.append(item)

        logger.info(f"關鍵字 '{keyword}' 列表頁找到 {len(items)} 筆工程類標案")
        return items

    async def parse_detail_page(self) -> dict:
        """解析詳細頁面，取得完整招標資訊"""
        page = self.page
        detail = {}

        await page.wait_for_load_state('networkidle')

        # 解析所有 table 中的 key-value 資料
        # 詳細頁的表格結構：th 是欄位名，td 是值
        rows = await page.query_selector_all('table tr')
        for row in rows:
            ths = await row.query_selector_all('th')
            tds = await row.query_selector_all('td')

            for i, th in enumerate(ths):
                key = (await th.inner_text()).strip().rstrip('：:')
                if i < len(tds):
                    # 取 td 的純文字，如果有 script 先處理
                    td = tds[i]
                    script = await td.query_selector('script')
                    if script:
                        script_text = await script.inner_html()
                        match = re.search(r'pageCode2Img\("([^"]+)"\)', script_text)
                        if match:
                            value = match.group(1)
                        else:
                            value = (await td.inner_text()).strip()
                    else:
                        value = (await td.inner_text()).strip()

                    if key and value:
                        detail[key] = value

        return detail

    async def scrape_detail(self, detail_url: str) -> dict:
        """訪問詳細頁（可能需要通過 CAPTCHA）"""
        page = self.page

        await page.goto(detail_url, wait_until='networkidle', timeout=30000)
        await asyncio.sleep(1)

        # 檢查是否遇到 CAPTCHA
        page_text = await page.text_content('body')
        if page_text and ('A區' in page_text or '撲克牌' in page_text or '防止惡意程式' in page_text):
            logger.info("遇到 CAPTCHA，開始辨識...")
            if await self.solve_captcha():
                await page.wait_for_load_state('networkidle')
                return await self.parse_detail_page()
            else:
                logger.error("CAPTCHA 解決失敗，跳過此標案")
                return {}
        else:
            # 沒有 CAPTCHA，直接解析
            return await self.parse_detail_page()

    async def scrape_keyword(self, keyword: str):
        """爬蟲單個關鍵字"""
        try:
            url = self.build_search_url(keyword)
            logger.info(f"爬蟲關鍵字: '{keyword}'")

            await self.page.goto(url, wait_until='networkidle', timeout=30000)
            await asyncio.sleep(2)

            # 解析列表頁
            items = await self.parse_list_page(keyword)

            for item in items:
                detail_url = item.pop('_detail_url', None)

                if detail_url:
                    try:
                        detail = await self.scrape_detail(detail_url)
                        # 合併詳細資料
                        item.update({
                            '機關地址': detail.get('機關地址', ''),
                            '聯絡人': detail.get('聯絡人', ''),
                            '聯絡電話': detail.get('聯絡電話', ''),
                            '傳真號碼': detail.get('傳真號碼', ''),
                            '採購金額級距': detail.get('採購金額級距', ''),
                            '是否適用條約或協定': detail.get('是否適用條約或協定之採購', ''),
                            '開標時間': detail.get('開標時間', ''),
                            '開標地點': detail.get('開標地點', ''),
                            '領標期限': detail.get('領標期限', ''),
                            '投標期限': detail.get('截止投標', detail.get('投標期限', '')),
                            '履約地點': detail.get('履約地點', ''),
                            '履約期限': detail.get('履約期限', ''),
                        })
                        await asyncio.sleep(1)  # 避免頻繁請求
                    except Exception as e:
                        logger.error(f"取得詳細頁失敗: {e}")

                item['抓取時間'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.data.append(item)

            logger.info(f"關鍵字 '{keyword}' 爬蟲完成，新增 {len(items)} 筆")

        except Exception as e:
            logger.error(f"爬蟲失敗 - 關鍵字: {keyword}, 錯誤: {e}")
            self.error_keywords.append({'關鍵字': keyword, '錯誤': str(e)})

    async def scrape_all(self):
        """爬蟲所有關鍵字"""
        start_date, end_date = self.get_date_range()
        logger.info("=" * 60)
        logger.info("開始爬蟲政府採購網招標資料 (Playwright 版)")
        logger.info(f"關鍵字數量: {len(KEYWORDS)}")
        logger.info(f"查詢日期範圍: {start_date} ~ {end_date}")
        logger.info("=" * 60)

        await self.start()

        try:
            for i, keyword in enumerate(KEYWORDS, 1):
                logger.info(f"進度: {i}/{len(KEYWORDS)}")
                await self.scrape_keyword(keyword)
                if i < len(KEYWORDS):
                    await asyncio.sleep(3)
        finally:
            await self.close()

        logger.info("=" * 60)
        logger.info(f"共爬蟲 {len(self.data)} 筆資料")
        if self.error_keywords:
            logger.warning(f"共 {len(self.error_keywords)} 個關鍵字出現錯誤")

    def export_to_excel(self):
        """匯出資料到 Excel"""
        if not self.data:
            logger.warning("沒有資料可匯出")
            return False

        try:
            output_dir = os.path.dirname(OUTPUT_FILE)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)

            df = pd.DataFrame(self.data)

            # 定義欄位順序
            col_order = [
                '關鍵字', '機關名稱', '標案案號', '標案名稱', '招標方式', '採購性質',
                '預算金額', '公告日期', '截止投標日期', '開標時間', '開標地點',
                '機關地址', '聯絡人', '聯絡電話', '傳真號碼',
                '採購金額級距', '領標期限', '投標期限', '履約地點', '履約期限',
                '是否適用條約或協定', '傳輸次數', '抓取時間'
            ]
            # 只保留存在的欄位
            existing_cols = [c for c in col_order if c in df.columns]
            extra_cols = [c for c in df.columns if c not in col_order]
            df = df[existing_cols + extra_cols]

            with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='招標資料', index=False)

                from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
                from openpyxl.utils import get_column_letter

                ws = writer.sheets['招標資料']

                thin_border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                header_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                header_font = Font(bold=True, color='000000')
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                max_lengths = {}
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), 1):
                    for col_idx, cell in enumerate(row, 1):
                        col_letter = get_column_letter(col_idx)
                        if cell.value:
                            cell_length = len(str(cell.value))
                            max_lengths[col_letter] = max(max_lengths.get(col_letter, 0), cell_length)
                        cell.border = thin_border
                        if row_idx == 1:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = header_alignment
                        else:
                            cell.alignment = Alignment(vertical='center', wrap_text=True)

                for col_letter, max_length in max_lengths.items():
                    ws.column_dimensions[col_letter].width = min(max_length * 1.5 + 2, 50)
                ws.row_dimensions[1].height = 30

            logger.info(f"資料已匯出到: {OUTPUT_FILE}")
            logger.info(f"共匯出 {len(self.data)} 筆資料")
            return True

        except Exception as e:
            logger.error(f"匯出 Excel 失敗: {e}")
            return False


def main():
    """主程序"""
    import argparse
    parser = argparse.ArgumentParser(description='政府採購網招標資料爬蟲 (Playwright)')
    parser.add_argument('--headless', action='store_true', help='無頭模式（不顯示瀏覽器）')
    parser.add_argument('--keywords', nargs='+', help='指定關鍵字（預設使用 config 中的全部關鍵字）')
    args = parser.parse_args()

    scraper = PlaywrightTenderScraper(headless=args.headless)

    if args.keywords:
        from config import KEYWORDS as _
        import config
        config.KEYWORDS = args.keywords

    try:
        asyncio.run(scraper.scrape_all())
        success = scraper.export_to_excel()
        if success:
            logger.info("爬蟲程序成功完成")
        else:
            logger.warning("爬蟲程序完成，但匯出失敗或無資料")

    except KeyboardInterrupt:
        logger.warning("程序被用戶中斷")
    except Exception as e:
        logger.error(f"程序執行出錯: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
