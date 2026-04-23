"""
政府採購網招標資料爬蟲
爬蟲政府採購網之招標資訊，並將結果匯出為 Excel 檔案
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime, timedelta
import json
from urllib.parse import urlencode
import time
import logging
from config import KEYWORDS, BASE_URL, HEADERS, REQUEST_TIMEOUT, OUTPUT_FILE

# 設定日誌
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class TenderScraper:
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update(HEADERS)
        self.data = []
        self.error_keywords = []
        
    def get_date_range(self):
        """取得當天日期範圍，格式為 YYYY/MM/DD"""
        end_date = datetime.now()
        start_date = end_date - timedelta(days=0)
        return start_date.strftime("%Y/%m/%d"), end_date.strftime("%Y/%m/%d")
    
    def build_url(self, keyword):
        """建構爬蟲URL"""
        start_date, end_date = self.get_date_range()
        
        params = {
            'firstSearch': 'true',
            'searchType': 'basic',
            'isBinding': 'N',
            'isLogIn': 'N',
            'orgName': '',
            'orgId': '',
            'tenderName': keyword,
            'tenderId': '',
            'tenderType': 'TENDER_DECLARATION',
            'tenderWay': 'TENDER_WAY_ALL_DECLARATION',
            'dateType': 'isNow',
            'tenderStartDate': start_date,
            'tenderEndDate': end_date,
            'radProctrgCate': '',
            'policyAdvocacy': ''
        }
        
        url = f"{BASE_URL}?{urlencode(params)}"
        return url
    
    def parse_response(self, html_content, keyword):
        """解析HTML回應"""
        soup = BeautifulSoup(html_content, 'lxml')
        
        try:
            # 查找主要資料表格 (class="tb_01" 或 id="tpam")
            table = soup.find('table', {'class': 'tb_01'}) or soup.find('table', {'id': 'tpam'})
            
            if not table:
                logger.info(f"關鍵字 '{keyword}' 未找到招標資料表格")
                return
            
            # 查找 tbody 中的數據行
            tbody = table.find('tbody')
            if not tbody:
                # 如果沒有 tbody，直接從 table 中查找
                rows = table.find_all('tr', {'class': 'tb_b2'})
            else:
                rows = tbody.find_all('tr')
            
            if not rows:
                logger.info(f"關鍵字 '{keyword}' 未找到招標資料行")
                return
            
            data_found = False
            
            # 處理每一行數據
            for row in rows:
                cols = row.find_all('td')
                if len(cols) >= 9:  # 至少要有 9 個欄位（項次到預算）
                    if self._extract_tender_info(cols, keyword):
                        data_found = True
            
            if data_found:
                logger.debug(f"關鍵字 '{keyword}' 成功提取 {len([d for d in self.data if d['關鍵字'] == keyword])} 筆資料")
            else:
                logger.info(f"關鍵字 '{keyword}' 未找到相符招標資料")
        
        except Exception as e:
            logger.error(f"解析HTML失敗 - 關鍵字: {keyword}, 錯誤: {e}")
    
    def _extract_tender_info(self, cols, keyword):
        """從欄位中提取招標資訊"""
        try:
            # 根據實際 HTML 結構提取資訊：
            # 第0欄：項次（序號）
            # 第1欄：機關名稱
            # 第2欄：標案案號 + 標案名稱
            # 第3欄：傳輸次數
            # 第4欄：招標方式
            # 第5欄：採購性質
            # 第6欄：公告日期
            # 第7欄：截止投標日期
            # 第8欄：預算金額
            
            tender_info = {
                '關鍵字': keyword,
                '抓取時間': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
            # 提取機關名稱（第1欄）
            if len(cols) > 1:
                tender_info['機關名稱'] = cols[1].get_text(strip=True)
            
            # 提取標案案號和標案名稱（第2欄）
            if len(cols) > 2:
                # 首先嘗試從 script 標籤中提取標案名稱
                script_tag = cols[2].find('script')
                tender_name_from_script = None
                
                if script_tag:
                    import re
                    script_text = script_tag.string
                    if script_text:
                        # 匹配 Geps3.CNS.pageCode2Img("標案名稱") 中的內容
                        match = re.search(r'Geps3\.CNS\.pageCode2Img\("([^"]+)"\)', script_text)
                        if match:
                            tender_name_from_script = match.group(1)
                
                # 提取標案案號
                col2_text = cols[2].get_text(separator='\n').strip()
                lines = [line.strip() for line in col2_text.split('\n') if line.strip()]
                
                if len(lines) >= 1:
                    tender_info['標案案號'] = lines[0]
                
                # 使用從 script 提取的名稱，如果沒有則使用文本提取
                if tender_name_from_script:
                    tender_info['標案名稱'] = tender_name_from_script
                elif len(lines) >= 2:
                    # 標案名稱可能包含多行，合併剩餘的行
                    tender_info['標案名稱'] = ' '.join(lines[1:])
            
            # 提取傳輸次數（第3欄）
            if len(cols) > 3:
                tender_info['傳輸次數'] = cols[3].get_text(strip=True)
            
            # 提取招標方式（第4欄）
            if len(cols) > 4:
                tender_info['招標方式'] = cols[4].get_text(strip=True)
            
            # 提取採購性質（第5欄）
            if len(cols) > 5:
                tender_info['採購性質'] = cols[5].get_text(strip=True)
            
            # 提取公告日期（第6欄）
            if len(cols) > 6:
                tender_info['公告日期'] = cols[6].get_text(strip=True)
            
            # 提取截止投標日期（第7欄）
            if len(cols) > 7:
                tender_info['截止投標日期'] = cols[7].get_text(strip=True)
            
            # 提取預算金額（第8欄）
            if len(cols) > 8:
                tender_info['預算金額'] = cols[8].get_text(strip=True)
            
            # 過濾條件：只保留採購性質為「工程」類的資料
            procurement_type = tender_info.get('採購性質', '')
            if '工程' not in procurement_type:
                return False
            
            # 只添加有標案案號或標案名稱的資訊
            if (tender_info.get('標案案號') and tender_info['標案案號']) or \
               (tender_info.get('標案名稱') and tender_info['標案名稱']):
                self.data.append(tender_info)
                return True
            
            return False
        
        except Exception as e:
            logger.error(f"解析行資料失敗: {e}")
            import traceback
            logger.debug(traceback.format_exc())
            return False
    
    def scrape_keyword(self, keyword):
        """爬蟲單個關鍵字"""
        url = self.build_url(keyword)
        
        try:
            logger.info(f"爬蟲關鍵字: '{keyword}'")
            
            response = self.session.get(url, timeout=REQUEST_TIMEOUT)
            response.raise_for_status()
            response.encoding = 'utf-8'
            
            self.parse_response(response.text, keyword)
            logger.info(f"關鍵字 '{keyword}' 爬蟲完成")
            
        except requests.exceptions.Timeout:
            error_msg = f"請求超時 - 關鍵字: {keyword}"
            logger.error(error_msg)
            self.error_keywords.append({'關鍵字': keyword, '錯誤': '請求超時'})
        except requests.exceptions.ConnectionError:
            error_msg = f"連接錯誤 - 關鍵字: {keyword}"
            logger.error(error_msg)
            self.error_keywords.append({'關鍵字': keyword, '錯誤': '連接錯誤'})
        except requests.exceptions.RequestException as e:
            error_msg = f"爬蟲失敗 - 關鍵字: {keyword}, 錯誤: {e}"
            logger.error(error_msg)
            self.error_keywords.append({'關鍵字': keyword, '錯誤': str(e)})
        except Exception as e:
            error_msg = f"未預期的錯誤 - 關鍵字: {keyword}, 錯誤: {e}"
            logger.error(error_msg)
            self.error_keywords.append({'關鍵字': keyword, '錯誤': str(e)})
    
    def scrape_all(self):
        """爬蟲所有關鍵字"""
        start_date, end_date = self.get_date_range()
        logger.info("=" * 60)
        logger.info("開始爬蟲政府採購網招標資料")
        logger.info(f"關鍵字數量: {len(KEYWORDS)}")
        logger.info(f"查詢日期範圍: {start_date} ~ {end_date}")
        logger.info("=" * 60)
        
        for i, keyword in enumerate(KEYWORDS, 1):
            logger.info(f"進度: {i}/{len(KEYWORDS)}")
            self.scrape_keyword(keyword)
            
            # 避免過於頻繁的請求
            if i < len(KEYWORDS):
                time.sleep(2)
        
        logger.info("=" * 60)
        logger.info(f"共爬蟲 {len(self.data)} 筆資料")
        if self.error_keywords:
            logger.warning(f"共 {len(self.error_keywords)} 個關鍵字出現錯誤")
    
    def export_to_excel(self):
        """匯出資料到Excel"""
        if not self.data:
            logger.warning("沒有資料可匯出")
            return False
        
        try:
            import os
            
            # 確保輸出目錄存在
            output_dir = os.path.dirname(OUTPUT_FILE)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            df = pd.DataFrame(self.data)
            
            # 使用 openpyxl 引擎
            with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='招標資料', index=False)
                
                # 導入 openpyxl 樣式模組
                from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
                from openpyxl.utils import get_column_letter
                
                ws = writer.sheets['招標資料']
                
                # 定義邊框樣式
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 定義標題行樣式（綠色背景）
                header_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                header_font = Font(bold=True, color='000000')
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # 計算最大列寬並應用樣式
                max_lengths = {}
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), 1):
                    for col_idx, cell in enumerate(row, 1):
                        col_letter = get_column_letter(col_idx)
                        
                        # 記錄最大長度
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if col_letter not in max_lengths:
                                max_lengths[col_letter] = cell_length
                            else:
                                max_lengths[col_letter] = max(max_lengths[col_letter], cell_length)
                        
                        # 應用邊框到所有單元格
                        cell.border = thin_border
                        
                        # 應用標題行樣式（第1行）
                        if row_idx == 1:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = header_alignment
                        else:
                            # 其他行垂直居中
                            cell.alignment = Alignment(vertical='center', wrap_text=True)
                
                # 設置列寬
                for col_letter, max_length in max_lengths.items():
                    adjusted_width = min(max_length * 1.5 + 2, 50)  # 最大寬度限制為50
                    ws.column_dimensions[col_letter].width = adjusted_width
                
                # 設置標題行高度
                ws.row_dimensions[1].height = 30
            
            logger.info(f"資料已匯出到: {OUTPUT_FILE}")
            logger.info(f"共匯出 {len(self.data)} 筆資料")
            return True
        
        except Exception as e:
            logger.error(f"匯出Excel失敗: {e}")
            return False
    
    def close(self):
        """關閉會話"""
        self.session.close()


def main():
    """主程序"""
    scraper = TenderScraper()
    
    try:
        # 爬蟲所有關鍵字
        scraper.scrape_all()
        
        # 匯出到Excel
        success = scraper.export_to_excel()
        
        if success:
            logger.info("爬蟲程序成功完成")
        else:
            logger.warning("爬蟲程序完成，但匯出失敗")
        
    except KeyboardInterrupt:
        logger.warning("程序被用戶中斷")
    except Exception as e:
        logger.error(f"程序執行出錯: {e}")
    finally:
        scraper.close()
        logger.info("爬蟲程序結束")


if __name__ == "__main__":
    main()
