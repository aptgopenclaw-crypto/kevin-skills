"""
政府採購網爬蟲 V2 - Playwright 版
完整流程：搜尋列表 → 撲克牌驗證碼 → 詳細頁 → Excel 報表
"""

import asyncio
import logging
import os
import random
import re
import time
from datetime import datetime, timedelta
from itertools import combinations
from urllib.parse import urlencode

import pandas as pd
from playwright.async_api import async_playwright, Page, BrowserContext

from config import (
    KEYWORDS, BASE_URL, HEADERS, OUTPUT_DIR, OUTPUT_FILE,
    REQUEST_DELAY, FILTER_PROCUREMENT_TYPE,
)
from detail_parser import parse_detail_page, extract_key_fields

logger = logging.getLogger(__name__)


class TenderScraperV2:
    def __init__(self, headless: bool = False):
        """
        Args:
            headless: 是否無頭模式（建議 False 以便觀察驗證碼）
        """
        self.headless = headless
        self.data = []           # 列表頁基本資料
        self.detail_data = []    # 詳細頁資料
        self.keyword_counts = {} # 每個關鍵字的筆數
        self.errors = []
        self.browser = None
        self.context = None

    async def init_browser(self):
        """初始化瀏覽器"""
        self.pw = await async_playwright().start()
        self.browser = await self.pw.chromium.launch(
            headless=self.headless,
            args=['--no-sandbox', '--disable-blink-features=AutomationControlled']
        )
        self.context = await self.browser.new_context(
            user_agent=HEADERS['User-Agent'],
            viewport={'width': 1280, 'height': 900},
            locale='zh-TW',
        )
        logger.info(f"瀏覽器已啟動 (headless={self.headless})")

    async def close(self):
        """關閉瀏覽器"""
        if self.context:
            await self.context.close()
        if self.browser:
            await self.browser.close()
        if self.pw:
            await self.pw.stop()
        logger.info("瀏覽器已關閉")

    def get_date_range(self):
        """取得當天日期範圍"""
        end_date = datetime.now()
        start_date = end_date - timedelta(days=0)
        return start_date.strftime("%Y/%m/%d"), end_date.strftime("%Y/%m/%d")

    def build_search_url(self, keyword: str) -> str:
        """建構搜尋 URL"""
        start_date, end_date = self.get_date_range()
        params = {
            'firstSearch': 'true',
            'searchType': 'basic',
            'isBinding': 'N',
            'isLogIn': 'N',
            'orgName': '',
            'orgId': '',
            'tenderName': keyword,
            'tenderId': '',
            'tenderType': 'TENDER_DECLARATION',
            'tenderWay': 'TENDER_WAY_ALL_DECLARATION',
            'dateType': 'isNow',
            'tenderStartDate': start_date,
            'tenderEndDate': end_date,
            'radProctrgCate': '',
            'policyAdvocacy': ''
        }
        return f"{BASE_URL}?{urlencode(params)}"

    async def scrape_list_page(self, page: Page, keyword: str) -> list[dict]:
        """爬取搜尋列表頁，回傳標案列表"""
        url = self.build_search_url(keyword)
        logger.info(f"搜尋關鍵字: '{keyword}'")

        try:
            await page.goto(url, wait_until='networkidle', timeout=30000)
            await page.wait_for_timeout(1000)
        except Exception as e:
            logger.error(f"載入列表頁失敗 ({keyword}): {e}")
            self.errors.append({'關鍵字': keyword, '錯誤': str(e)})
            return []

        # 解析列表表格 — 用 Playwright DOM API 直接取，同時提取 detail URL
        rows = await page.query_selector_all('table.tb_01 tbody tr, table#tpam tbody tr, tr.tb_b2')
        tenders = []

        # DEBUG: dump 列表頁第一筆 row 的完整 HTML 及整頁所有 <a> 含 href 的內容
        if rows:
            debug_dir = os.path.join(OUTPUT_DIR, "debug")
            os.makedirs(debug_dir, exist_ok=True)
            # dump first row HTML
            first_row_html = await rows[0].inner_html()
            debug_file = os.path.join(debug_dir, f"list_row_debug_{keyword}.html")
            with open(debug_file, 'w', encoding='utf-8') as f:
                f.write(first_row_html)
            logger.info(f"  DEBUG: 已儲存第一筆 row HTML → {debug_file}")

            # dump 所有 row 裡的 <a> 標籤
            for ri, r in enumerate(rows[:3]):
                links = await r.query_selector_all('a')
                for li, link in enumerate(links):
                    href = await link.get_attribute('href') or ''
                    onclick = await link.get_attribute('onclick') or ''
                    text = (await link.inner_text()).strip()
                    logger.info(f"  DEBUG row[{ri}] a[{li}]: text='{text[:30]}' href='{href[:80]}' onclick='{onclick[:80]}'")

        for row in rows:
            cols = await row.query_selector_all('td')
            if len(cols) < 9:
                continue

            # 提取 detail URL — 從「檢視」連結取得
            # 實際格式: /prkms/urlSelector/common/tpam?pk=BASE64_ID
            detail_url = ''
            view_link = await row.query_selector('a[href*="tpam?pk="], a[href*="urlSelector"]')
            if not view_link:
                # fallback: 找 text=檢視 的連結
                view_link = await row.query_selector('a:has-text("檢視")')
            if view_link:
                href = await view_link.get_attribute('href') or ''
                if href:
                    detail_url = href if href.startswith('http') else f"https://web.pcc.gov.tw{href}"

            # 取各欄位文字
            col_texts = []
            for col in cols:
                col_texts.append((await col.inner_text()).strip())

            # 標案案號/名稱可能有 script 編碼
            col2_text = col_texts[2] if len(col_texts) > 2 else ''
            script_el = await cols[2].query_selector('script') if len(cols) > 2 else None
            tender_name_from_script = ''
            if script_el:
                script_text = await script_el.inner_html()
                match = re.search(r'pageCode2Img\("([^"]+)"\)', script_text)
                if match:
                    tender_name_from_script = match.group(1)

            tenders.append({
                '項次': col_texts[0] if len(col_texts) > 0 else '',
                '機關名稱': col_texts[1] if len(col_texts) > 1 else '',
                '標案案號_名稱': col2_text,
                '_標案名稱_script': tender_name_from_script,
                '傳輸次數': col_texts[3] if len(col_texts) > 3 else '',
                '招標方式': col_texts[4] if len(col_texts) > 4 else '',
                '採購性質': col_texts[5] if len(col_texts) > 5 else '',
                '公告日期': col_texts[6] if len(col_texts) > 6 else '',
                '截止投標': col_texts[7] if len(col_texts) > 7 else '',
                '預算金額': col_texts[8] if len(col_texts) > 8 else '',
                '_detail_url': detail_url,
            })

        # 過濾採購性質
        filtered = []
        for t in tenders:
            if FILTER_PROCUREMENT_TYPE and FILTER_PROCUREMENT_TYPE not in t.get('採購性質', ''):
                continue
            t['關鍵字'] = keyword
            # 拆分案號和名稱
            parts = t.get('標案案號_名稱', '').split('\n')
            parts = [p.strip() for p in parts if p.strip()]
            t['標案案號'] = parts[0] if parts else ''
            # 優先用 script 取得的名稱
            if t.get('_標案名稱_script'):
                t['標案名稱'] = t['_標案名稱_script']
            else:
                t['標案名稱'] = ' '.join(parts[1:]) if len(parts) > 1 else ''
            filtered.append(t)

        logger.info(f"  關鍵字 '{keyword}': 找到 {len(tenders)} 筆，過濾後 {len(filtered)} 筆")
        return filtered

    async def solve_captcha_and_get_detail(self, page: Page, tender: dict) -> dict | None:
        """
        直接 navigate 到 detail URL → 處理撲克牌驗證碼（窮舉法）→ 取得詳細頁資料

        流程：
        1. page.goto(detail_url) → 出現撲克牌 CAPTCHA 整頁
        2. 窮舉 C(6,2)=15 種組合嘗試配對
        3. 點擊 B區匹配的牌 → 確認送出
        4. 成功後頁面跳轉到詳細頁 → 解析

        Args:
            page: Playwright Page
            tender: 列表頁的標案資料 dict（需含 _detail_url）

        Returns:
            詳細頁解析後的 dict，或 None
        """
        detail_url = tender.get('_detail_url', '')
        case_no = tender.get('標案案號', '')
        org_name = tender.get('機關名稱', '')
        logger.info(f"  嘗試取得詳細資料: {org_name} / {case_no}")

        if not detail_url:
            logger.warning(f"  沒有 detail URL: {case_no}")
            return None

        # 前往 detail URL（會先出現 CAPTCHA 頁）
        try:
            await page.goto(detail_url, wait_until='networkidle', timeout=30000)
            await page.wait_for_timeout(1000)
        except Exception as e:
            logger.error(f"  載入 detail URL 失敗: {e}")
            return None

        # 檢查是否需要 CAPTCHA
        page_text = await page.text_content('body') or ''
        if '機關資料' in page_text or '招標資料' in page_text or '機關代碼' in page_text:
            # 不需要 CAPTCHA，直接到詳細頁
            logger.info(f"  直接進入詳細頁（無驗證碼）")
            html = await page.content()
            return parse_detail_page(html)

        # 有 CAPTCHA — 開始解碼流程
        has_captcha = ('A區' in page_text or '撲克牌' in page_text or
                       '防止惡意程式' in page_text or '驗證' in page_text)
        if not has_captcha:
            logger.warning(f"  頁面既非詳細頁也非 CAPTCHA: {page.url}")
            # 儲存截圖 debug
            debug_dir = os.path.join(OUTPUT_DIR, "debug")
            os.makedirs(debug_dir, exist_ok=True)
            debug_path = os.path.join(debug_dir, f"unknown_{case_no}.png")
            await page.screenshot(path=debug_path)
            logger.info(f"  未知頁面截圖: {debug_path}")
            return None

        # ── 隨機窮舉法：每次隨機選 C(6,2) 中的一種組合，直到成功 ──
        # 伺服器每次提交失敗後會更換驗證碼，每次命中率 1/15
        all_combos = list(combinations(range(6), 2))
        attempt = 0

        while True:
            attempt += 1
            combo = random.choice(all_combos)
            logger.info(f"  窮舉法嘗試 {attempt}，隨機選擇: 索引 {combo}")

            # 取得 B區 checkboxes
            b_checkboxes = await page.query_selector_all('input[name="choose"]')
            if len(b_checkboxes) < 6:
                logger.warning(f"  B區牌數不足: {len(b_checkboxes)}，嘗試重新整理")
                refresh_btn = await page.query_selector('#b_refresh, text=重新整理, text=重新產生')
                if refresh_btn:
                    await refresh_btn.click()
                    await page.wait_for_timeout(2000)
                continue

            # 透過 JavaScript 清除所有 checkbox 狀態
            for cb in b_checkboxes:
                await cb.evaluate('el => el.checked = false')

            # 點擊選中的兩張牌的 label
            for idx in combo:
                cb_id = await b_checkboxes[idx].get_attribute('id')
                if cb_id:
                    label = await page.query_selector(f'label[for="{cb_id}"]')
                    if label:
                        await label.click()
                        await page.wait_for_timeout(300)
                    else:
                        await b_checkboxes[idx].evaluate('el => el.checked = true')
                else:
                    await b_checkboxes[idx].evaluate('el => el.checked = true')
                logger.info(f"  已選擇 B區第 {idx} 張牌")

            # 點擊「確認送出」
            submit_btn = await page.query_selector('#b_submit, input[value="確認送出"]')
            if not submit_btn:
                submit_btn = await page.query_selector('button:has-text("送出")')
            if submit_btn:
                await submit_btn.click()
                logger.info(f"  已點擊確認送出")
            else:
                logger.warning(f"  找不到確認送出按鈕")
                continue

            # 等待頁面回應
            await page.wait_for_timeout(3000)
            try:
                await page.wait_for_load_state('networkidle', timeout=10000)
            except Exception:
                pass

            # 檢查是否驗證成功
            new_text = await page.text_content('body') or ''

            if '機關資料' in new_text or '招標資料' in new_text or '機關代碼' in new_text:
                logger.info(f"  ✓ 窮舉法驗證碼通過！組合: {combo}，嘗試次數: {attempt}")
                html = await page.content()
                return parse_detail_page(html)

            # 還在 CAPTCHA 頁 = 該組合不對，繼續
            if 'A區' in new_text or '撲克牌' in new_text or '驗證' in new_text:
                logger.info(f"  組合 {combo} 未通過，繼續嘗試...")
                continue

            # 未知狀態
            logger.warning(f"  未知驗證結果，URL: {page.url}")

    async def scrape_all(self):
        """主流程：爬取所有關鍵字"""
        await self.init_browser()
        page = await self.context.new_page()

        start_date, end_date = self.get_date_range()
        logger.info("=" * 60)
        logger.info("政府採購網招標資料爬蟲 V2（含詳細頁）")
        logger.info(f"關鍵字數量: {len(KEYWORDS)}")
        logger.info(f"查詢日期: {start_date} ~ {end_date}")
        logger.info(f"過濾採購性質: {FILTER_PROCUREMENT_TYPE or '不過濾'}")
        logger.info("=" * 60)

        all_tenders = []

        for i, keyword in enumerate(KEYWORDS, 1):
            logger.info(f"進度: {i}/{len(KEYWORDS)}")
            tenders = await self.scrape_list_page(page, keyword)
            self.keyword_counts[keyword] = len(tenders)

            for j, tender in enumerate(tenders):
                tender['_row_idx'] = j
                all_tenders.append(tender)

            if i < len(KEYWORDS):
                await page.wait_for_timeout(REQUEST_DELAY * 1000)

        self.data = all_tenders
        logger.info(f"列表頁共取得 {len(all_tenders)} 筆標案")

        # 取得每筆標案的詳細資料
        if all_tenders:
            logger.info("=" * 60)
            logger.info("開始取得詳細頁資料...")
            logger.info("=" * 60)

            for i, tender in enumerate(all_tenders, 1):
                logger.info(f"詳細頁進度: {i}/{len(all_tenders)}")

                # 直接用 detail URL 前往（不需要重新載入列表頁）
                detail = await self.solve_captcha_and_get_detail(page, tender)
                if detail:
                    key_fields = extract_key_fields(detail)
                    # 合併列表頁和詳細頁資料
                    merged = {**tender, **key_fields, '_detail_raw': detail}
                    self.detail_data.append(merged)
                else:
                    # 沒有詳細頁，只用列表頁資料
                    self.detail_data.append(tender)

                await page.wait_for_timeout(REQUEST_DELAY * 1000)

        logger.info("=" * 60)
        logger.info(f"爬蟲完成: 列表 {len(self.data)} 筆, 詳細 {len(self.detail_data)} 筆")
        if self.errors:
            logger.warning(f"錯誤: {len(self.errors)} 個關鍵字")

    def export_to_excel(self):
        """匯出 Excel（PCC 含連結版格式）"""
        data_to_export = self.detail_data if self.detail_data else self.data
        if not data_to_export:
            logger.warning("沒有資料可匯出")
            return False

        try:
            os.makedirs(OUTPUT_DIR, exist_ok=True)

            # 整理匯出欄位（11 欄，對齊範本格式）
            export_rows = []
            detail_urls = []  # 保存 detail URL 供後續加超連結
            for item in data_to_export:
                detail_url = item.get('_detail_url', '')
                row = {
                    '項次': item.get('關鍵字', ''),
                    '機關名稱': item.get('機關名稱', ''),
                    '標案案號': item.get('標案案號', ''),
                    '標案名稱': item.get('標案名稱', ''),
                    '傳輸次數': item.get('傳輸次數', ''),
                    '招標方式': item.get('招標方式', ''),
                    '採購性質': item.get('採購性質', ''),
                    '公告日期': item.get('公告日期', item.get('公告日', '')),
                    '截止投標': item.get('截止投標', ''),
                    '預算金額': item.get('預算金額', ''),
                    '詳細連結': detail_url,
                }
                export_rows.append(row)
                detail_urls.append(detail_url)

            df = pd.DataFrame(export_rows)

            with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='招標資料', index=False)
                self._style_excel(writer, detail_urls)

            logger.info(f"已匯出: {OUTPUT_FILE} ({len(export_rows)} 筆)")
            return True

        except Exception as e:
            logger.error(f"匯出 Excel 失敗: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _style_excel(self, writer, detail_urls):
        """Excel 樣式（PCC 含連結版格式）"""
        from openpyxl.styles import Border, Side, Font, Alignment
        from openpyxl.worksheet.hyperlink import Hyperlink

        ws = writer.sheets['招標資料']

        # 範本欄寬
        col_widths = {
            'A': 18,    # 項次
            'B': 31,    # 機關名稱
            'C': 22,    # 標案案號
            'D': 72,    # 標案名稱
            'E': 13,    # 傳輸次數
            'F': 13,    # 招標方式
            'G': 13,    # 採購性質
            'H': 13,    # 公告日期
            'I': 13,    # 截止投標
            'J': 13.5,  # 預算金額
            'K': 13,    # 詳細連結
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # 字體定義
        base_font = Font(name='新細明體', size=12)
        header_font = Font(name='新細明體', size=12, bold=True)
        link_font = Font(name='新細明體', size=12, underline='single', color='0563C1')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_align = Alignment(vertical='center', wrap_text=False)

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), 1):
            for col_idx, cell in enumerate(row, 1):
                cell.border = thin_border
                if row_idx == 1:
                    cell.font = header_font
                    cell.alignment = header_align
                else:
                    cell.font = base_font
                    cell.alignment = cell_align

        # 標案名稱 (D 欄) 加超連結
        for data_row_idx, url in enumerate(detail_urls):
            if url:
                cell = ws.cell(row=data_row_idx + 2, column=4)  # D 欄, +2 跳過 header
                cell.hyperlink = Hyperlink(ref=cell.coordinate, target=url)
                cell.font = link_font

        ws.row_dimensions[1].height = 30


async def main():
    """主入口"""
    import argparse
    parser = argparse.ArgumentParser(description='政府採購網爬蟲 V2')
    parser.add_argument('--headless', action='store_true', help='無頭模式')
    parser.add_argument('--list-only', action='store_true', help='只爬列表頁，不爬詳細頁')
    args = parser.parse_args()

    scraper = TenderScraperV2(
        headless=args.headless,
    )

    try:
        if args.list_only:
            await scraper.init_browser()
            page = await scraper.context.new_page()
            for keyword in KEYWORDS:
                tenders = await scraper.scrape_list_page(page, keyword)
                scraper.keyword_counts[keyword] = len(tenders)
                scraper.data.extend(tenders)
                await page.wait_for_timeout(REQUEST_DELAY * 1000)
            scraper.detail_data = scraper.data
        else:
            await scraper.scrape_all()

        scraper.export_to_excel()

        # 寄送郵件
        if not args.list_only:
            from mail_sender import send_report_email
            send_report_email(
                excel_path=OUTPUT_FILE,
                keyword_counts=scraper.keyword_counts,
                total=len(scraper.detail_data),
            )

    except KeyboardInterrupt:
        logger.warning("使用者中斷")
    except Exception as e:
        logger.error(f"執行錯誤: {e}")
        import traceback
        traceback.print_exc()
    finally:
        await scraper.close()


if __name__ == '__main__':
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    asyncio.run(main())
