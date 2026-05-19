"""
政府採購網爬蟲 V3 配置

關鍵字設定由 keywords_config.xlsx 維護（位於 v3/ 根目錄）。
首次執行時會自動產生該檔案，使用者可直接在 Excel 中新增 / 修改 / 刪除，
儲存後重新執行程式即生效，無需修改此 Python 檔案。
"""
from datetime import date
import os
import logging
from collections import OrderedDict

# ─── 預設關鍵字（工程師維護的基準值） ──────────────────────────────────────────

_DEFAULT_SOLUTION_KEYWORD_MAP = OrderedDict([
    ("智慧三表", ["瓦斯表", "水表", "AMI", "電表"]),
    ("共杆", ["節點", "樞紐", "共桿", "共杆"]),
    ("路燈", ["路燈"]),
    ("水情", ["洩洪", "水情", "下水道"]),
    ("空氣品質", ["空品", "空氣品質"]),
    ("班班（中小學）", ["新風", "冷氣", "EMS"]),
    ("充電樁", ["充電", "停車"]),
    ("交通－站牌", ["公車", "站牌"]),
    ("交通－號誌不斷電", ["不斷電"]),
    ("交通－ITS", ["車路協同", "號誌", "交控", "行車安全", "路口", "自動化駕駛", "交通", "人流", "車流"]),
    ("CCTV", ["執法", "監視系統", "辨識", "偵測", "測速", "車牌", "取締", "違規", "錄影"]),
    ("ESG－淨零及節能", ["溫室", "能效", "冰水主機", "碳排放", "低碳", "能源", "淨零", "節能"]),
    ("ESG－微電網及儲能", ["儲能", "電網", "太陽", "再生能源"]),
    ("IoT相關", ["聯線", "連線", "通訊", "AI", "人工智慧", "監控", "5G"]),
])

_DEFAULT_SOLUTION_ORG_KEYWORD_MAP = {
    "交通－ITS": ["政府", "交通", "警察", "公路局", "園區"],
    "班班（中小學）": ["國民中學", "小學", "高級中學", "職業學校"],
    "交通－號誌不斷電": ["政府", "交通", "警察", "公路局", "園區"],
    "ESG－建研所補助": ["警察專科學校", "考選部", "海巡署中部分署", "農業部水產試驗所",
                     "國立花蓮高級商業學校", "國軍退除役官兵輔導委員會", "國立金門大學"],
}

_DEFAULT_ORG_ONLY_SOLUTIONS = ["ESG－建研所補助"]

# ─── Excel 設定檔路徑（v3/ 根目錄） ───────────────────────────────────────────

_CONFIG_EXCEL = os.path.join(os.path.dirname(os.path.dirname(__file__)), "keywords_config.xlsx")


def _create_excel_template(path: str) -> None:
    """建立初始的 keywords_config.xlsx，供使用者自行維護關鍵字。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    # ── Sheet 1: 關鍵字設定 ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "關鍵字設定"
    ws1.append(["Solution名稱", "關鍵字"])
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for sol, kws in _DEFAULT_SOLUTION_KEYWORD_MAP.items():
        for kw in kws:
            ws1.append([sol, kw])

    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 20

    ws1.append([])
    ws1.append(["【說明】"])
    ws1.append(["1. 同一個 Solution 可以有多個關鍵字，每列填一個"])
    ws1.append(["2. 新增 Solution：在 Solution名稱 欄填入新名稱，關鍵字欄填入對應關鍵字"])
    ws1.append(["3. 刪除關鍵字：直接刪除該列"])
    ws1.append(["4. 存檔後重新執行程式即生效，無需修改程式碼"])

    # ── Sheet 2: 機關過濾 ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("機關過濾")
    ws2.append(["Solution名稱", "機關關鍵字", "僅機關搜尋"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for sol, org_kws in _DEFAULT_SOLUTION_ORG_KEYWORD_MAP.items():
        is_org_only = "是" if sol in _DEFAULT_ORG_ONLY_SOLUTIONS else "否"
        for org_kw in org_kws:
            ws2.append([sol, org_kw, is_org_only])

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 12

    ws2.append([])
    ws2.append(["【說明】"])
    ws2.append(["1. 機關關鍵字：搜尋結果的機關名稱需包含此關鍵字才納入報表"])
    ws2.append(["2. 僅機關搜尋（是/否）：填「是」時改以機關名稱逐一搜尋，不使用標案名稱關鍵字"])
    ws2.append(["3. 不需要機關過濾的 Solution 不必在此表填寫"])

    wb.save(path)
    logging.info(f"已自動產生關鍵字設定檔：{path}，請依需求修改後重新執行。")


def _load_from_excel(path: str):
    """從 Excel 讀取設定，回傳 (solution_kw_map, solution_org_map, org_only_solutions)。
    任一步驟失敗時回傳 (None, None, None)，由呼叫端 fallback 到預設值。"""
    import pandas as pd

    try:
        df_kw = pd.read_excel(path, sheet_name="關鍵字設定", dtype=str)
        df_org = pd.read_excel(path, sheet_name="機關過濾", dtype=str)
    except Exception as e:
        logging.warning(f"讀取 {path} 失敗，使用預設關鍵字：{e}")
        return None, None, None

    def _is_valid(val: str) -> bool:
        return bool(val) and val != "nan" and not val.startswith("【")

    # 建立 SOLUTION_KEYWORD_MAP（保留列順序）
    solution_kw_map: OrderedDict = OrderedDict()
    for _, row in df_kw.iterrows():
        sol = str(row.get("Solution名稱", "")).strip()
        kw = str(row.get("關鍵字", "")).strip()
        if _is_valid(sol) and _is_valid(kw):
            solution_kw_map.setdefault(sol, []).append(kw)

    if not solution_kw_map:
        logging.warning("keywords_config.xlsx 的「關鍵字設定」工作表為空，使用預設關鍵字。")
        return None, None, None

    # 建立 SOLUTION_ORG_KEYWORD_MAP 和 ORG_ONLY_SOLUTIONS
    solution_org_map: dict = {}
    org_only_solutions: list = []
    for _, row in df_org.iterrows():
        sol = str(row.get("Solution名稱", "")).strip()
        kw = str(row.get("機關關鍵字", "")).strip()
        only = str(row.get("僅機關搜尋", "否")).strip()
        if _is_valid(sol) and _is_valid(kw):
            solution_org_map.setdefault(sol, []).append(kw)
            if only == "是" and sol not in org_only_solutions:
                org_only_solutions.append(sol)

    total_kw = sum(len(v) for v in solution_kw_map.values())
    logging.info(
        f"從 keywords_config.xlsx 載入 {len(solution_kw_map)} 個 Solution、"
        f"{total_kw} 個關鍵字。"
    )
    return solution_kw_map, solution_org_map, org_only_solutions


# ─── 載入設定 ──────────────────────────────────────────────────────────────────

if not os.path.exists(_CONFIG_EXCEL):
    _create_excel_template(_CONFIG_EXCEL)
    SOLUTION_KEYWORD_MAP = _DEFAULT_SOLUTION_KEYWORD_MAP
    SOLUTION_ORG_KEYWORD_MAP = _DEFAULT_SOLUTION_ORG_KEYWORD_MAP
    ORG_ONLY_SOLUTIONS = _DEFAULT_ORG_ONLY_SOLUTIONS
else:
    _kw_map, _org_map, _org_only = _load_from_excel(_CONFIG_EXCEL)
    SOLUTION_KEYWORD_MAP = _kw_map or _DEFAULT_SOLUTION_KEYWORD_MAP
    SOLUTION_ORG_KEYWORD_MAP = _org_map if _org_map is not None else _DEFAULT_SOLUTION_ORG_KEYWORD_MAP
    ORG_ONLY_SOLUTIONS = _org_only if _org_only is not None else _DEFAULT_ORG_ONLY_SOLUTIONS

# 從 SOLUTION_KEYWORD_MAP 自動產生關鍵字清單（去重保序）
KEYWORDS = []
for _kws in SOLUTION_KEYWORD_MAP.values():
    for _kw in _kws:
        if _kw not in KEYWORDS:
            KEYWORDS.append(_kw)

# 關鍵字 → Solution 反查表
KEYWORD_TO_SOLUTION = {}
for _sol, _kws in SOLUTION_KEYWORD_MAP.items():
    for _kw in _kws:
        KEYWORD_TO_SOLUTION[_kw] = _sol

# 基礎 URL
BASE_URL = "https://web.pcc.gov.tw/prkms/tender/common/basic/readTenderBasic"
DETAIL_URL = "https://web.pcc.gov.tw/tps/QueryTender/query/searchTenderDetail"

# HTTP 請求標頭
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36'
}

# 請求超時時間（秒）
REQUEST_TIMEOUT = 30

# 輸出目錄
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "output")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"{date.today().strftime('%Y%m%d')}招標資料.xlsx")

# 請求間隔（秒）
REQUEST_DELAY = 2

# 過濾：採購性質（空清單 = 不過濾）
FILTER_PROCUREMENT_TYPE = ["工程", "財物", "勞務"]

# Email 設定（從環境變數讀取，可在此覆蓋）
MAIL_SMTP_HOST = os.environ.get("O_MAILSER", "smtp.office365.com")
MAIL_SMTP_PORT = int(os.environ.get("O_MAILPORT", "587"))
MAIL_USER = os.environ.get("O_MAILUSER", "fetprivatenetwork@feto365.tw")
MAIL_PASSWORD = os.environ.get("O_MAILPASS", "FET5gpns")
MAIL_ALIAS = os.environ.get("O_ALIAS", "政府採購公告")
MAIL_RECIPIENTS = [
    "kevinchang4@fareastone.com.tw",
]
