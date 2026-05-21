"""
政府採購網決標資料爬蟲
流程：搜尋決標列表 → 撲克牌驗證碼（C(6,2) 窮舉法）→ 詳細頁 → Excel 報表 → 寄信
"""

import asyncio
import logging
import os
import random
import re
from datetime import datetime, timedelta
from itertools import combinations
from urllib.parse import urlencode

import pandas as pd
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright, Page

from config import (
    KEYWORDS, HEADERS, OUTPUT_DIR,
    REQUEST_DELAY, FILTER_PROCUREMENT_TYPE, KEYWORD_TO_SOLUTION,
    SOLUTION_ORG_KEYWORD_MAP, ORG_ONLY_SOLUTIONS,
)

logger = logging.getLogger(__name__)

AWARD_BASE_URL = "https://web.pcc.gov.tw/prkms/tender/common/agent/readTenderAgent"
AWARD_DETAIL_BASE = "https://web.pcc.gov.tw/tps/atm/AtmAwardWithoutSso/QueryAtmAwardDetail"
AWARD_OUTPUT_FILE = os.path.join(
    OUTPUT_DIR,
    f"{datetime.now().strftime('%Y%m%d')}決標資料.xlsx",
)


# ─── 詳細頁解析 ────────────────────────────────────────────────────────────────

def _clean(text: str) -> str:
    if not text:
        return ""
    return re.sub(r"\s+", " ", text).strip()


def parse_award_detail(html: str) -> dict:
    """解析決標公告詳細頁 HTML，回傳結構化 dict。"""
    soup = BeautifulSoup(html, "lxml")
    result = {}

    # ── 機關資料（tbg_1 / tbg_2）────────────────────────────────────────────
    _parse_label_value_rows(soup, "tbg_1", "tbg_2", result)

    # ── 採購資料（tbg_4 / tbg_4R）───────────────────────────────────────────
    _parse_label_value_rows(soup, "tbg_4", "tbg_4R", result)

    # ── 投標廠商（tbg_6 / tbg_4R）───────────────────────────────────────────
    _parse_bidders(soup, result)

    # ── 決標品項（tbg_7 / tbg_4R）───────────────────────────────────────────
    _parse_label_value_rows(soup, "tbg_7", "tbg_4R", result)

    # ── 決標資料（tbg_9 / tbg_4R）───────────────────────────────────────────
    _parse_label_value_rows(soup, "tbg_9", "tbg_4R", result)

    # ── 數值欄位（hidden input）──────────────────────────────────────────────
    for input_id, field_name in [
        ("budgetAmount", "預算金額"),
        ("governmentEstimate", "底價金額"),
    ]:
        el = soup.find("input", {"id": input_id})
        if el and el.get("value"):
            result.setdefault(field_name, el["value"])

    return result


def _parse_label_value_rows(soup: BeautifulSoup, label_cls: str, value_cls: str, result: dict):
    """通用：依 label/value class 解析所有 tr 配對。"""
    for row in soup.find_all("tr"):
        label_cell = row.find("td", class_=label_cls)
        value_cell = row.find("td", class_=value_cls)
        if label_cell and value_cell:
            label = _clean(label_cell.get_text())
            value = _clean(value_cell.get_text())
            if label and value and label not in result:
                result[label] = value


def _parse_bidders(soup: BeautifulSoup, result: dict):
    """解析投標廠商區塊，提取得標廠商主要欄位。"""
    # 投標廠商家數（直接文字節點，用 get_text 比對）
    all_tbg6 = soup.find_all("td", class_="tbg_6")
    for td in all_tbg6:
        if "投標廠商家數" in td.get_text():
            val_td = td.find_next_sibling("td")
            if val_td:
                result["投標廠商家數"] = _clean(val_td.get_text())
            break

    # 找所有廠商區塊：「投標廠商N」的 td 開始
    # 實際 HTML 中文字包在巢狀 span/div 裡，不能用 string= 直接比對
    bidders = []
    vendor_rows = [td for td in all_tbg6 if re.search(r"投標廠商\d+", td.get_text())]
    for vendor_td in vendor_rows:
        bidder = {}
        # 從「投標廠商N」的 tr 往下爬
        row = vendor_td.find_parent("tr")
        while row:
            row = row.find_next_sibling("tr")
            if not row:
                break
            label_td = row.find("td", class_="tbg_6")
            value_td = row.find("td", class_="tbg_4R")
            if not label_td or not value_td:
                # 遇到下一家廠商的起始行就停
                if row.find("td", class_="tbg_6", string=re.compile(r"投標廠商\d+")):
                    break
                continue
            label = _clean(label_td.get_text()).lstrip("　")
            value = _clean(value_td.get_text())
            if label and value:
                bidder[label] = value
            # 遇到「履約起迄日期」或「雇用員工」通常是該廠商最後一行
            if label in ("履約起迄日期", "雇用員工總人數是否超過100人"):
                # 還會有更多廠商後，繼續找下一個 vendor_td
                break
        if bidder:
            bidders.append(bidder)

    result["_bidders"] = bidders

    # 整理第一家「得標廠商」的主要欄位到頂層
    winner = next(
        (b for b in bidders if b.get("是否得標") == "是"),
        bidders[0] if bidders else None,
    )
    if winner:
        result["得標廠商名稱"] = winner.get("廠商名稱", "")
        result["得標廠商代碼"] = winner.get("廠商代碼", "")
        result["得標廠商決標金額"] = winner.get("決標金額", "")
        result["得標廠商國別"] = winner.get("得標廠商國別", "")
        result["是否為中小企業"] = winner.get("是否為中小企業", "")
        result["履約起迄日期"] = winner.get("履約起迄日期", "")


# ─── 爬蟲主體 ──────────────────────────────────────────────────────────────────

class AwardScraper:
    """政府採購網決標資料爬蟲（Playwright 版）"""

    def __init__(self, headless: bool = False):
        self.headless = headless
        self.data: list[dict] = []
        self.detail_data: list[dict] = []
        self.keyword_counts: dict[str, int] = {}
        self.errors: list[dict] = []
        self.browser = None
        self.context = None
        self.pw = None

    async def init_browser(self):
        self.pw = await async_playwright().start()
        self.browser = await self.pw.chromium.launch(
            headless=self.headless,
            args=["--no-sandbox", "--disable-blink-features=AutomationControlled"],
        )
        self.context = await self.browser.new_context(
            user_agent=HEADERS["User-Agent"],
            viewport={"width": 1280, "height": 900},
            locale="zh-TW",
        )
        logger.info(f"瀏覽器已啟動 (headless={self.headless})")

    async def close(self):
        if self.context:
            await self.context.close()
        if self.browser:
            await self.browser.close()
        if self.pw:
            await self.pw.stop()
        logger.info("瀏覽器已關閉")

    def get_date_range(self):
        """回傳今天的日期字串（起迄相同）"""
        today = datetime.now().strftime("%Y/%m/%d")
        return today, today

    def build_search_url(self, keyword: str = "", org_name: str = "") -> str:
        """建構決標搜尋 URL"""
        start_date, end_date = self.get_date_range()
        params = {
            "firstSearch": "true",
            "isQuery": "Y",
            "isBinding": "N",
            "isLogIn": "N",
            "orgName": org_name,
            "orgId": "",
            "tenderName": keyword,
            "tenderId": "",
            "tenderStatus": "TENDER_STATUS_1",
            "tenderWay": "TENDER_WAY_ALL_DECLARATION",
            "awardAnnounceStartDate": start_date,
            "awardAnnounceEndDate": end_date,
            "radProctrgCate": "",
            "tenderRange": "TENDER_RANGE_ALL",
            "minBudget": "",
            "maxBudget": "",
            "item": "",
            "gottenVendorName": "",
            "gottenVendorId": "",
            "submitVendorName": "",
            "submitVendorId": "",
            "execLocation": "",
            "priorityCate": "",
            "radReConstruct": "",
            "policyAdvocacy": "",
            "isCpp": "",
            "pageSize": "100",
        }
        return f"{AWARD_BASE_URL}?{urlencode(params)}"

    async def scrape_list_page(
        self, page: Page, keyword: str = "", org_name: str = "", solution: str = ""
    ) -> list[dict]:
        """爬取決標搜尋列表頁，回傳標案列表（含決標公告 detail URL）"""
        url = self.build_search_url(keyword=keyword, org_name=org_name)
        search_label = org_name or keyword or "(全部)"
        logger.info(f"  決標搜尋: '{search_label}'")

        try:
            await page.goto(url, wait_until="networkidle", timeout=30000)
            await page.wait_for_timeout(1000)
        except Exception as e:
            logger.error(f"  載入決標列表頁失敗 ({search_label}): {e}")
            self.errors.append({"關鍵字": search_label, "錯誤": str(e)})
            return []

        # 檢查是否需要 CAPTCHA（列表頁也可能有）
        page_text = await page.text_content("body") or ""
        if "A區" in page_text or "撲克牌" in page_text or "防止惡意程式" in page_text:
            logger.info("  列表頁出現驗證碼，開始解碼...")
            success = await self._solve_captcha(page)
            if not success:
                logger.error("  列表頁驗證碼失敗，跳過此關鍵字")
                return []
            await page.wait_for_timeout(1000)

        # 解析結果表格 table#atm
        rows = await page.query_selector_all("table#atm tbody tr.tntr, table#atm tbody tr")
        tenders = []
        _debug_saved = False  # 每次搜尋只存一次 debug HTML

        for row in rows:
            cols = await row.query_selector_all("td")
            if len(cols) < 8:
                continue

            col_texts = [_clean(await col.inner_text()) for col in cols]
            # 跳過「無符合條件資料」行
            if "無符合條件" in col_texts[0] or not col_texts[0].strip():
                continue

            # DEBUG：儲存第一筆 row 的 HTML，方便診斷連結格式
            if not _debug_saved:
                debug_dir = os.path.join(OUTPUT_DIR, "debug")
                os.makedirs(debug_dir, exist_ok=True)
                row_html_debug = await row.inner_html()
                with open(os.path.join(debug_dir, f"award_list_row_{search_label}.html"), "w", encoding="utf-8") as _f:
                    _f.write(row_html_debug)
                _debug_saved = True

            # 取得決標公告詳細頁 URL — 多層次提取
            detail_url = await self._extract_detail_url_from_row(row, cols)

            # 拆分 標案案號/名稱（col[2]：多行文字）
            case_raw = col_texts[2] if len(col_texts) > 2 else ""
            parts = [p for p in case_raw.split("\n") if p.strip()]
            case_no = parts[0].strip() if parts else ""
            case_name = " ".join(p.strip() for p in parts[1:]) if len(parts) > 1 else ""

            tenders.append({
                "項次": col_texts[0] if col_texts else "",
                "機關名稱": col_texts[1] if len(col_texts) > 1 else "",
                "標案案號": case_no,
                "標案名稱": case_name,
                "招標方式": col_texts[3] if len(col_texts) > 3 else "",
                "標的分類": col_texts[4] if len(col_texts) > 4 else "",
                "公告日期": col_texts[5] if len(col_texts) > 5 else "",
                "決標金額_列表": col_texts[6] if len(col_texts) > 6 else "",
                "決標公告序號": col_texts[7] if len(col_texts) > 7 else "",
                "無法決標": col_texts[8] if len(col_texts) > 8 else "",
                "_detail_url": detail_url,
                "_list_url": url,          # 備用：回列表頁點擊時使用
                "_case_no": case_no,
                "關鍵字": keyword or org_name,
                "_solution": solution or KEYWORD_TO_SOLUTION.get(keyword, ""),
            })

        # 採購性質過濾（以標的分類為準）
        if FILTER_PROCUREMENT_TYPE:
            before = len(tenders)
            tenders = [
                t for t in tenders
                if any(pt in t.get("標的分類", "") for pt in FILTER_PROCUREMENT_TYPE)
            ]
            logger.info(f"  採購性質過濾: {before} → {len(tenders)} 筆")

        # 機關名稱過濾（同招標爬蟲邏輯）
        sol = solution or KEYWORD_TO_SOLUTION.get(keyword, "")
        if keyword and not org_name:
            org_keywords = SOLUTION_ORG_KEYWORD_MAP.get(sol, [])
            if org_keywords:
                before = len(tenders)
                tenders = [t for t in tenders if any(ok in t.get("機關名稱", "") for ok in org_keywords)]
                logger.info(f"  機關名稱過濾: {before} → {len(tenders)} 筆")

        # 統計 detail_url 找到的比率
        found = sum(1 for t in tenders if t.get("_detail_url"))
        logger.info(f"  '{search_label}': 決標 {len(tenders)} 筆，detail URL 取得 {found}/{len(tenders)} 筆")
        return tenders

    async def _extract_detail_url_from_row(self, row, cols) -> str:
        """
        從列表頁 row 提取決標詳細頁 URL。

        實際 HTML 中連結格式為：
          https://web.pcc.gov.tw/prkms/urlSelector/common/atm?pk=BASE64_ID
        其中 pk 值等同於 QueryAtmAwardDetail 的 pkAtmMain 值。
        """
        row_html = await row.inner_html()

        # ── 方法1：最常見 — urlSelector/common/atm?pk= ────────────────────────
        # 列表頁的決標公告、標案名稱、檢視按鈕都用這個格式
        m = re.search(r'urlSelector/common/atm\?pk=([A-Za-z0-9+/=]+)', row_html)
        if m:
            pk = m.group(1)
            logger.debug(f"  method1 pk={pk}")
            return f"{AWARD_DETAIL_BASE}?pkAtmMain={pk}"

        # ── 方法2：直接含 QueryAtmAwardDetail（未來若格式變更）────────────────
        m = re.search(r'href=["\']([^"\']*QueryAtmAwardDetail[^"\']*)["\']', row_html)
        if m:
            path = m.group(1)
            return path if path.startswith("http") else f"https://web.pcc.gov.tw{path}"

        # ── 方法3：任意位置含 pkAtmMain= ─────────────────────────────────────
        m = re.search(r'pkAtmMain=([A-Za-z0-9+/=]{8,})', row_html)
        if m:
            return f"{AWARD_DETAIL_BASE}?pkAtmMain={m.group(1).rstrip(chr(39) + chr(34))}"

        # ── 方法4：透過 JS evaluate 取得決標公告欄(col[7])連結的 href ────────
        if len(cols) > 7:
            link = await cols[7].query_selector("a")
            if not link and len(cols) > 9:
                link = await cols[9].query_selector("a")   # 功能選項欄「檢視」
            if link:
                try:
                    href = await link.get_attribute("href") or ""
                    m4 = re.search(r'pk=([A-Za-z0-9+/=]+)', href)
                    if m4:
                        return f"{AWARD_DETAIL_BASE}?pkAtmMain={m4.group(1)}"
                except Exception:
                    pass

        logger.debug(f"  detail URL 未找到，row HTML (前300): {row_html[:300]}")
        return ""

    async def _solve_captcha(self, page: Page) -> bool:
        """撲克牌驗證碼窮舉法（同 scraper_v2.py 邏輯）"""
        all_combos = list(combinations(range(6), 2))
        attempt = 0
        while True:
            attempt += 1
            combo = random.choice(all_combos)
            logger.info(f"  驗證碼嘗試 {attempt}，組合: {combo}")

            b_checkboxes = await page.query_selector_all('input[name="choose"]')
            if len(b_checkboxes) < 6:
                logger.warning(f"  B區牌數不足: {len(b_checkboxes)}")
                refresh_btn = await page.query_selector("#b_refresh, text=重新整理, text=重新產生")
                if refresh_btn:
                    await refresh_btn.click()
                    await page.wait_for_timeout(2000)
                continue

            for cb in b_checkboxes:
                await cb.evaluate("el => el.checked = false")

            for idx in combo:
                cb_id = await b_checkboxes[idx].get_attribute("id")
                if cb_id:
                    label = await page.query_selector(f'label[for="{cb_id}"]')
                    if label:
                        await label.click()
                        await page.wait_for_timeout(300)
                    else:
                        await b_checkboxes[idx].evaluate("el => el.checked = true")
                else:
                    await b_checkboxes[idx].evaluate("el => el.checked = true")

            submit_btn = await page.query_selector('#b_submit, input[value="確認送出"]')
            if not submit_btn:
                submit_btn = await page.query_selector('button:has-text("送出")')
            if submit_btn:
                await submit_btn.click()
            else:
                continue

            await page.wait_for_timeout(3000)
            try:
                await page.wait_for_load_state("networkidle", timeout=10000)
            except Exception:
                pass

            new_text = await page.text_content("body") or ""

            # 成功條件：出現決標詳細頁的標誌性內容
            if "機關代碼" in new_text or "決標日期" in new_text or "標案案號" in new_text:
                logger.info(f"  ✓ 驗證碼通過，嘗試: {attempt}")
                return True

            # 仍在驗證碼頁：繼續嘗試下一組合
            if "A區" in new_text or "撲克牌" in new_text or "驗證" in new_text:
                logger.info(f"  組合 {combo} 未通過，繼續嘗試...")

            if attempt >= 30:
                logger.error("  驗證碼嘗試 30 次仍失敗")
                return False

    async def get_award_detail(self, page: Page, tender: dict) -> dict | None:
        """前往決標詳細頁取得完整資料（QueryAtmAwardDetail 為公開頁，通常無 CAPTCHA）"""
        detail_url = tender.get("_detail_url", "")
        case_no = tender.get("標案案號", "")
        org_name = tender.get("機關名稱", "")

        # ── fallback：detail URL 未取得時，回列表頁點擊 ──────────────────────
        if not detail_url:
            logger.info(f"  detail URL 未取得，嘗試點擊取得: {org_name}/{case_no}")
            detail_url = await self._get_detail_url_by_click(page, tender)
            if detail_url:
                tender["_detail_url"] = detail_url   # 回填供日後參考
            else:
                logger.warning(f"  仍無法取得 detail URL: {org_name}/{case_no}")
                return None

        logger.info(f"  取得決標詳細: {org_name} / {case_no}")
        try:
            await page.goto(detail_url, wait_until="networkidle", timeout=30000)
            await page.wait_for_timeout(800)
        except Exception as e:
            logger.error(f"  載入決標詳細頁失敗: {e}")
            return None

        page_text = await page.text_content("body") or ""

        # 有 CAPTCHA？
        if "A區" in page_text or "撲克牌" in page_text or "防止惡意程式" in page_text:
            logger.info("  決標詳細頁出現驗證碼，開始解碼...")
            success = await self._solve_captcha(page)
            if not success:
                return None
            await page.wait_for_timeout(1000)

        html = await page.content()
        # 確認是詳細頁內容
        if "機關代碼" not in html and "決標日期" not in html:
            logger.warning(f"  詳細頁內容異常: {page.url}")
            debug_dir = os.path.join(OUTPUT_DIR, "debug")
            os.makedirs(debug_dir, exist_ok=True)
            await page.screenshot(path=os.path.join(debug_dir, f"award_detail_{case_no}.png"))
            return None

        return parse_award_detail(html)

    async def _get_detail_url_by_click(self, page: Page, tender: dict) -> str:
        """
        Fallback：回到列表頁，找到對應 row，點擊「決標公告」連結，
        攔截 navigation 取得 detail URL，再回到列表頁。
        """
        list_url = tender.get("_list_url", "")
        case_no = tender.get("_case_no", tender.get("標案案號", ""))

        if not list_url:
            return ""

        try:
            await page.goto(list_url, wait_until="networkidle", timeout=30000)
            await page.wait_for_timeout(1000)
        except Exception as e:
            logger.warning(f"  回列表頁失敗: {e}")
            return ""

        # 找到案號相符的 row
        rows = await page.query_selector_all("table#atm tbody tr")
        target_row = None
        for row in rows:
            cols = await row.query_selector_all("td")
            if len(cols) < 8:
                continue
            cell_text = await cols[2].inner_text() if len(cols) > 2 else ""
            if case_no and case_no in cell_text:
                target_row = row
                break
            # 備選：機關名稱
            if not case_no:
                target_row = row
                break

        if not target_row:
            logger.warning(f"  列表頁找不到案號 '{case_no}'")
            return ""

        # 在 決標公告欄（col[7]）找連結並點擊，攔截 navigation
        target_cols = await target_row.query_selector_all("td")
        link_to_click = None
        if len(target_cols) > 7:
            link_to_click = await target_cols[7].query_selector("a")
        if not link_to_click:
            link_to_click = await target_row.query_selector("a")
        if not link_to_click:
            return ""

        captured: list[str] = []

        async def _on_request(req):
            if "QueryAtmAwardDetail" in req.url or "pkAtmMain" in req.url:
                captured.append(req.url)

        page.on("request", _on_request)
        old_url = page.url
        try:
            await link_to_click.click()
            await page.wait_for_timeout(2000)
            try:
                await page.wait_for_load_state("networkidle", timeout=8000)
            except Exception:
                pass
        except Exception as e:
            logger.warning(f"  點擊決標連結失敗: {e}")
        finally:
            page.remove_listener("request", _on_request)

        found_url = ""
        if captured:
            found_url = captured[0]
        elif page.url != old_url and "QueryAtmAwardDetail" in page.url:
            found_url = page.url

        if found_url:
            logger.info(f"  ✓ 點擊取得 detail URL: {found_url}")
            # 導回列表頁（若已離開）
            if page.url != list_url:
                try:
                    await page.go_back()
                    await page.wait_for_timeout(1000)
                except Exception:
                    pass
        return found_url

    async def scrape_all(self):
        """主流程：依 config.py 關鍵字爬取所有決標資料"""
        await self.init_browser()
        page = await self.context.new_page()

        start_date, end_date = self.get_date_range()
        logger.info("=" * 60)
        logger.info("政府採購網決標資料爬蟲")
        logger.info(f"關鍵字數量: {len(KEYWORDS)}")
        logger.info(f"決標公告日期: {start_date} ~ {end_date}")
        logger.info("=" * 60)

        all_tenders: list[dict] = []

        # 關鍵字搜尋
        for i, keyword in enumerate(KEYWORDS, 1):
            logger.info(f"關鍵字進度: {i}/{len(KEYWORDS)}")
            tenders = await self.scrape_list_page(page, keyword=keyword)
            self.keyword_counts[keyword] = len(tenders)
            all_tenders.extend(tenders)
            if i < len(KEYWORDS):
                await page.wait_for_timeout(REQUEST_DELAY * 1000)

        # 機關名稱搜尋（org-only solutions）
        for sol in ORG_ONLY_SOLUTIONS:
            org_names = SOLUTION_ORG_KEYWORD_MAP.get(sol, [])
            logger.info(f"機關名稱搜尋 [{sol}]：{len(org_names)} 個機關")
            for i_org, org in enumerate(org_names, 1):
                logger.info(f"  機關 {i_org}/{len(org_names)}: {org}")
                tenders = await self.scrape_list_page(page, org_name=org, solution=sol)
                self.keyword_counts[org] = len(tenders)
                all_tenders.extend(tenders)
                await page.wait_for_timeout(REQUEST_DELAY * 1000)

        self.data = all_tenders
        logger.info(f"列表頁共取得 {len(all_tenders)} 筆決標資料")

        # 取得詳細頁資料
        if all_tenders:
            logger.info("=" * 60)
            logger.info("開始取得決標詳細頁資料...")
            logger.info("=" * 60)
            for i, tender in enumerate(all_tenders, 1):
                logger.info(f"詳細頁進度: {i}/{len(all_tenders)}")
                detail = await self.get_award_detail(page, tender)
                if detail:
                    merged = {**tender, **detail}
                    # 以詳細頁機關名稱為準（覆蓋列表頁）
                    if detail.get("機關名稱"):
                        merged["機關名稱"] = detail["機關名稱"]
                    if detail.get("標案名稱"):
                        merged["標案名稱"] = detail["標案名稱"]
                    self.detail_data.append(merged)
                else:
                    self.detail_data.append(tender)
                await page.wait_for_timeout(REQUEST_DELAY * 1000)

        logger.info("=" * 60)
        logger.info(f"爬蟲完成: 列表 {len(self.data)} 筆，詳細 {len(self.detail_data)} 筆")

    def export_to_excel(self) -> bool:
        """匯出決標 Excel"""
        data_to_export = self.detail_data if self.detail_data else self.data
        if not data_to_export:
            logger.warning("沒有決標資料可匯出")
            return False

        os.makedirs(OUTPUT_DIR, exist_ok=True)
        export_rows = []

        for idx, item in enumerate(data_to_export, 1):
            keyword = item.get("關鍵字", "")
            solution = item.get("_solution") or KEYWORD_TO_SOLUTION.get(keyword, "")
            row = {
                "項次": idx,
                "相關的Solution": solution,
                "關鍵字": keyword,
                "機關名稱": item.get("機關名稱", ""),
                "標案案號": item.get("標案案號", ""),
                "標案名稱": item.get("標案名稱", ""),
                "招標方式": item.get("招標方式", ""),
                "標的分類": item.get("標的分類", ""),
                # 決標公告日期：優先用詳細頁
                "決標公告日期": item.get("決標公告日期", item.get("公告日期", "")),
                "決標日期": item.get("決標日期", ""),
                "預算金額": item.get("預算金額", ""),
                "總決標金額": item.get("總決標金額", item.get("決標金額_列表", "")),
                "底價金額": item.get("底價金額", ""),
                "採購金額級距": item.get("採購金額級距", ""),
                "決標方式": item.get("決標方式", ""),
                "決標資料類別": item.get("決標資料類別", ""),
                "是否複數決標": item.get("是否複數決標", ""),
                "投標廠商家數": item.get("投標廠商家數", ""),
                "得標廠商名稱": item.get("得標廠商名稱", ""),
                "得標廠商代碼": item.get("得標廠商代碼", ""),
                "得標廠商決標金額": item.get("得標廠商決標金額", ""),
                "得標廠商國別": item.get("得標廠商國別", ""),
                "是否為中小企業": item.get("是否為中小企業", ""),
                "履約起迄日期": item.get("履約起迄日期", ""),
                # 機關聯絡資訊
                "機關代碼": item.get("機關代碼", ""),
                "單位名稱": item.get("單位名稱", ""),
                "機關地址": item.get("機關地址", ""),
                "聯絡人": item.get("聯絡人", ""),
                "聯絡電話": item.get("聯絡電話", ""),
                "傳真號碼": item.get("傳真號碼", ""),
                "辦理方式": item.get("辦理方式", ""),
                "履約地點": item.get("履約地點", ""),
                "契約編號": item.get("契約編號", ""),
                "是否刊登公報": item.get("是否刊登公報", ""),
                "詳細連結": item.get("_detail_url", ""),
            }
            export_rows.append(row)

        try:
            df = pd.DataFrame(export_rows)
            with pd.ExcelWriter(AWARD_OUTPUT_FILE, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="決標資料", index=False)
                self._style_excel(writer)
            logger.info(f"已匯出: {AWARD_OUTPUT_FILE} ({len(export_rows)} 筆)")
            return True
        except Exception as e:
            logger.error(f"匯出 Excel 失敗: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _style_excel(self, writer):
        """Excel 樣式設定"""
        from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        ws = writer.sheets["決標資料"]
        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        link_font = Font(color="0563C1", underline="single")

        header_row = [cell.value for cell in ws[1]]
        name_col = (header_row.index("標案名稱") + 1) if "標案名稱" in header_row else None
        url_col = (header_row.index("詳細連結") + 1) if "詳細連結" in header_row else None

        max_lengths: dict[str, int] = {}
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), 1
        ):
            for col_idx, cell in enumerate(row, 1):
                col_letter = get_column_letter(col_idx)
                if cell.value:
                    max_lengths[col_letter] = max(max_lengths.get(col_letter, 0), len(str(cell.value)))
                cell.border = thin
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                else:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

        # 標案名稱加超連結
        if name_col and url_col:
            for row_idx in range(2, ws.max_row + 1):
                url_cell = ws.cell(row=row_idx, column=url_col)
                name_cell = ws.cell(row=row_idx, column=name_col)
                if url_cell.value and name_cell.value:
                    name_cell.hyperlink = str(url_cell.value)
                    name_cell.font = link_font

        for col_letter, max_len in max_lengths.items():
            ws.column_dimensions[col_letter].width = min(max_len * 1.5 + 2, 50)
        ws.row_dimensions[1].height = 30


async def main():
    """決標爬蟲主入口"""
    import argparse

    parser = argparse.ArgumentParser(description="政府採購網決標資料爬蟲")
    parser.add_argument("--headless", action="store_true", help="無頭模式")
    parser.add_argument("--list-only", action="store_true", help="只爬列表頁，不爬詳細頁")
    args = parser.parse_args()

    scraper = AwardScraper(headless=args.headless)

    try:
        if args.list_only:
            await scraper.init_browser()
            page = await scraper.context.new_page()
            from config import KEYWORDS as ALL_KW
            for keyword in ALL_KW:
                tenders = await scraper.scrape_list_page(page, keyword=keyword)
                scraper.keyword_counts[keyword] = len(tenders)
                scraper.data.extend(tenders)
                await page.wait_for_timeout(REQUEST_DELAY * 1000)
            scraper.detail_data = scraper.data
        else:
            await scraper.scrape_all()

        scraper.export_to_excel()

        if not args.list_only:
            from mail_sender import send_award_report_email
            send_award_report_email(
                excel_path=AWARD_OUTPUT_FILE,
                keyword_counts=scraper.keyword_counts,
                total=len(scraper.detail_data),
            )

    except KeyboardInterrupt:
        logger.warning("使用者中斷")
    except Exception as e:
        logger.error(f"執行錯誤: {e}")
        import traceback
        traceback.print_exc()
    finally:
        await scraper.close()


if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
    )
    asyncio.run(main())
